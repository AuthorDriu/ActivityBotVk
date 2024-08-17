from typing import List
from tempfile import NamedTemporaryFile
from datetime import datetime as dt
from datetime import UTC

import openpyxl as oxl
from openpyxl.styles import PatternFill

from schemas import Result


# Возвращает файл в байтовом виде
def generate_xl(results: List[Result], date: str) -> bytes:
    # Создание книги Excel
    wb = oxl.Workbook()
    # Получение активного листа
    ws = wb.active
    
    # Заголовки столбцов
    ws.cell(row=1, column=1).value = "URL пользователя"
    ws.cell(row=1, column=2).value = "Продолжительность активности"
    ws.cell(row=1, column=3).value = "Отклонение от нормы"
    
    # Заполнение листа данными
    for i, result in enumerate(results, start=2):
        cell_user_url = ws.cell(row=i, column=1)
        cell_user_active_time = ws.cell(row=i, column=2)
        cell_user_difference = ws.cell(row=i, column=3)
        
        cell_user_url.value = result.user.to_url()
        cell_user_active_time.value = _format_time(abs(result.active_time_utc), "%H:%M:%S")
        cell_user_difference.value = _format_time(abs(result.difference_utc), "%H:%M:%S")
        
        # Окрашивание ячеек
        if result.difference_utc > 0: # Красный цвет, если пользователь был активен меньше положенного
            cell_user_difference.fill = PatternFill(start_color="FA8072", end_color="FA8072", fill_type="solid")
        else: # Зелёный, если был активен положенное время или
            cell_user_difference.fill = PatternFill(start_color="7CFC00", end_color="7CFC00", fill_type="solid")

    # Сохранение в виде потока
    with NamedTemporaryFile(mode="rb") as tmp:
        wb.save(tmp.name)
        wb.close()
        f = tmp.read()
        return f


def _format_time(utc_timestamp: dt, format: str) -> str:
    utc = dt.fromtimestamp(utc_timestamp, UTC)
    return utc.strftime(format)
        